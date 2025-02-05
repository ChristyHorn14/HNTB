import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the input Excel file
input_file = '/Users/chrishornung/Desktop/HNTB Reboot/Active Tumor Board LINKED.xlsx'
sheet_name = 'Master Linked'
sorted_df = pd.read_excel(input_file, sheet_name=sheet_name)


# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Define the date style
date_style = NamedStyle(name="date_style", number_format='MM/DD/YYYY')

# Function to convert a cell to datetime if possible
def convert_to_date(cell_value):
    try:
        return pd.to_datetime(cell_value)
    except:
        return cell_value

# Convert relevant columns to datetime
date_columns = ['DOB', 'HP/Clinic', 'Imaging1', 'Imaging 2', 'Imaging 3', 'OR1', 'OR2', 'OR3', 'Path1', 'Path2', 'Path3']
for col in date_columns:
    sorted_df[col] = sorted_df[col].apply(convert_to_date)

# Function to write patient data
def write_patient_data(patient_data, ws, start_row):
    # Writing patient's basic details
    ws[f'A{start_row}'] = patient_data['Last name']
    ws[f'B{start_row}'] = patient_data['First name']
    ws[f'C{start_row}'] = 'DOB'
    if pd.notnull(patient_data['DOB']) and isinstance(patient_data['DOB'], pd.Timestamp):
        ws[f'D{start_row}'] = patient_data['DOB'].strftime('%m/%d/%Y')
        ws[f'D{start_row}'].style = date_style  # Apply date style to DOB cell
    ws[f'E{start_row}'] = 'MRN'
    ws[f'F{start_row}'] = patient_data['EPIC MRN']
    ws[f'G{start_row}'] = 'Attending'
    ws[f'H{start_row}'] = patient_data['Attending']
    
    # Move to the next row for the summary
    current_row = start_row + 1
    ws[f'A{current_row}'] = patient_data['Summary']
    
    # Writing patient's clinic details
    current_row += 1
    ws[f'A{current_row}'] = 'HP/Clinic'
    if pd.notnull(patient_data['HP/Clinic']) and isinstance(patient_data['HP/Clinic'], pd.Timestamp):
        ws[f'B{current_row}'] = patient_data['HP/Clinic'].strftime('%m/%d/%Y')
        ws[f'B{current_row}'].style = date_style  # Apply date style to HP/Clinic cell
    ws[f'E{current_row}'] = 'List'
    ws[f'F{current_row}'] = patient_data['List']

    # Writing patient's imaging details
    current_row += 1
    ws[f'A{current_row}'] = 'Imaging1'
    if pd.notnull(patient_data['Imaging1']) and isinstance(patient_data['Imaging1'], pd.Timestamp):
        ws[f'B{current_row}'] = patient_data['Imaging1'].strftime('%m/%d/%Y')
        ws[f'B{current_row}'].style = date_style  # Apply date style to Imaging1 cell
    ws[f'C{current_row}'] = 'Imaging2'
    if pd.notnull(patient_data['Imaging 2']) and isinstance(patient_data['Imaging 2'], pd.Timestamp):
        ws[f'D{current_row}'] = patient_data['Imaging 2'].strftime('%m/%d/%Y')
        ws[f'D{current_row}'].style = date_style  # Apply date style to Imaging2 cell
    ws[f'E{current_row}'] = 'Imaging3'
    if pd.notnull(patient_data['Imaging 3']) and isinstance(patient_data['Imaging 3'], pd.Timestamp):
        ws[f'F{current_row}'] = patient_data['Imaging 3'].strftime('%m/%d/%Y')
        ws[f'F{current_row}'].style = date_style  # Apply date style to Imaging3 cell
    
    # Writing patient's OR details
    current_row += 1
    ws[f'A{current_row}'] = 'OR1'
    if pd.notnull(patient_data['OR1']) and isinstance(patient_data['OR1'], pd.Timestamp):
        ws[f'B{current_row}'] = patient_data['OR1'].strftime('%m/%d/%Y')
        ws[f'B{current_row}'].style = date_style  # Apply date style to OR1 cell
    ws[f'C{current_row}'] = 'OR2'
    if pd.notnull(patient_data['OR2']) and isinstance(patient_data['OR2'], pd.Timestamp):
        ws[f'D{current_row}'] = patient_data['OR2'].strftime('%m/%d/%Y')
        ws[f'D{current_row}'].style = date_style  # Apply date style to OR2 cell
    ws[f'E{current_row}'] = 'OR3'
    if pd.notnull(patient_data['OR3']) and isinstance(patient_data['OR3'], pd.Timestamp):
        ws[f'F{current_row}'] = patient_data['OR3'].strftime('%m/%d/%Y')
        ws[f'F{current_row}'].style = date_style  # Apply date style to OR3 cell
    
    # Writing patient's pathology details
    current_row += 1
    ws[f'A{current_row}'] = 'Path1'
    if pd.notnull(patient_data['Path1']) and isinstance(patient_data['Path1'], pd.Timestamp):
        ws[f'B{current_row}'] = patient_data['Path1'].strftime('%m/%d/%Y')
        ws[f'B{current_row}'].style = date_style  # Apply date style to Path1 cell
    ws[f'C{current_row}'] = 'Path2'
    if pd.notnull(patient_data['Path2']) and isinstance(patient_data['Path2'], pd.Timestamp):
        ws[f'D{current_row}'] = patient_data['Path2'].strftime('%m/%d/%Y')
        ws[f'D{current_row}'].style = date_style  # Apply date style to Path2 cell
    ws[f'E{current_row}'] = 'Path3'
    if pd.notnull(patient_data['Path3']) and isinstance(patient_data['Path3'], pd.Timestamp):
        ws[f'F{current_row}'] = patient_data['Path3'].strftime('%m/%d/%Y')
        ws[f'F{current_row}'].style = date_style  # Apply date style to Path3 cell
    ws[f'G{current_row}'] = 'Resident'
    ws[f'H{current_row}'] = patient_data['Resident']

    # Writing patient's other notes
    current_row += 1
    ws[f'A{current_row}'] = 'Other notes'
    ws[f'B{current_row}'] = patient_data['Other Notes']

    # Adjust start_row for the next patient
    return current_row + 2  # Add 2 rows to separate each patient's data

# Iterate over each row in the DataFrame and write the data
start_row = 1
for index, row in sorted_df.iterrows():
    start_row = write_patient_data(row, ws, start_row)

# Set the column widths as specified (approximation to character widths)
column_widths = {
    'A': 25,
    'B': 50,
    'C': 17,
    'D': 50,
    'E': 16,
    'F': 41,
    'G': 19,
    'H': 25,
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook to a file
print("file saved successfully")
output_file = '/Users/chrishornung/Desktop/HNTB Reboot/Outputs/face_sheets.xlsx'
wb.save(output_file)
